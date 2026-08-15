import json
import shutil
import subprocess
import sys
import urllib.request
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

timesheets_path = Path(__file__).parent / "data" / "timesheets"
stackexchange_path = Path(__file__).parent / "data" / "stackexchange"
geodata_path = Path(__file__).parent / "data" / "geodata"
projectstatus_path = Path(__file__).parent / "data" / "projectstatus"

# The Excel files are generated from sample.json and are rebuilt every time - that costs a second.
# The downloads are about 15 MB from three different sites, most of it countries.geojson, so they
# are skipped when the files are already there. To fetch them again:
#
#   python 05_sample_data_setup.py --force
force = "--force" in sys.argv


def download(url, target):
    # Download to a temporary name and rename only once the whole file has arrived. A half written
    # countries.geojson otherwise stays behind looking like a perfectly good file, and fails much
    # later inside demo 3 as a JSONDecodeError that says nothing about a download.
    temporary = target.with_name(target.name + ".part")

    # noqa on the next line, not in ruff.toml: every url passed in here is a literal further down
    # this file, so `file:` and the other schemes S310 warns about cannot get in. It stays a local
    # exemption so that a download added later, with a url built from a variable, still trips it.
    with urllib.request.urlopen(url, timeout=60) as response, temporary.open("wb") as file:  # noqa: S310
        shutil.copyfileobj(response, file)
        expected = response.headers.get("Content-Length")

    # A server that closes the connection early gives us a short file and no error at all, so
    # compare what arrived against what was announced. Chunked responses announce nothing.
    size = temporary.stat().st_size
    if expected is not None and size != int(expected):
        temporary.unlink()
        raise OSError(f"{target.name} is incomplete: got {size} bytes, expected {expected}")

    temporary.replace(target)


# TimeSheets
# Excel files will be generated from sample.json
print("Setting up Excel files for TimeSheets")

sample_data = json.loads((timesheets_path / "sample.json").read_text(encoding="utf-8"))

# One workbook per department, one worksheet per person inside it
departments = sorted({row["Department"] for row in sample_data})

for department in departments:
    department_data = [row for row in sample_data if row["Department"] == department]
    persons = sorted({row["Person"] for row in department_data})

    workbook = Workbook()
    workbook.remove(workbook.active)

    for person in persons:
        person_data = [row for row in department_data if row["Person"] == person]
        worksheet = workbook.create_sheet(title=person)

        # The note that the employees are supposed to read
        worksheet["A1"] = "Please fill out this form weekly and send it to HR. Thanks!"
        worksheet["A1"].font = Font(size=14, bold=True)

        # The header row - the importer skips the first two rows and reads this one
        for column, header in enumerate(["date", "time_from", "time_to", "project", "task"], start=1):
            cell = worksheet.cell(row=3, column=column, value=header)
            cell.font = Font(bold=True)
            if column <= 3:
                cell.alignment = Alignment(horizontal="right")

        # One row per booking, starting right below the header
        for index, row in enumerate(person_data, start=4):
            start = datetime.fromisoformat(row["Start"])
            end = datetime.fromisoformat(row["End"])

            worksheet.cell(row=index, column=1, value=start.date()).number_format = "dd.mm.yyyy"
            worksheet.cell(row=index, column=2, value=start.time()).number_format = "HH:mm"
            worksheet.cell(row=index, column=3, value=end.time()).number_format = "HH:mm"
            worksheet.cell(row=index, column=4, value=row["Project"])
            worksheet.cell(row=index, column=5, value=row["Task"])

        # Make the columns wide enough to read
        for column, width in {"A": 12, "B": 12, "C": 12, "D": 20, "E": 20}.items():
            worksheet.column_dimensions[column].width = width

    file = timesheets_path / f"{department}.xlsx"
    workbook.save(file)
    print(f"Created {file.name} with {len(persons)} worksheets and {len(department_data)} rows")


# StackExchange
# XML files will be downloaded from archive.org/download/stackexchange
site = "dba.meta"
archive = stackexchange_path / "tmp.7z"
xml_files = sorted(stackexchange_path.glob("*.xml"))

if xml_files and not force:
    print(f"Keeping the {len(xml_files)} StackExchange XML files that are already there")
else:
    print(f"Downloading StackExchange data for {site}")

    download(f"https://archive.org/download/stackexchange/{site}.stackexchange.com.7z", archive)

    print(f"Downloaded {archive.name} with {archive.stat().st_size / 1024 / 1024:.1f} MB")

    # 7za is installed by 02_wsl2_setup.sh, the same way the sibling repository uses it.
    # "e" extracts all files of the archive into the working directory.
    subprocess.run(["7za", "e", "-y", archive.name], cwd=stackexchange_path, check=True, capture_output=True)

    archive.unlink()

    for file in sorted(stackexchange_path.glob("*.xml")):
        print(f"Created {file.name} with {file.stat().st_size / 1024 / 1024:.1f} MB")


# Geodata
# GPX files will be downloaded from https://www.berlin.de/sen/uvk/mobilitaet-und-verkehr/verkehrsplanung/radverkehr/radverkehrsnetz/radrouten/gpx/
# One GPX file will be downloaded from https://www.michael-mueller-verlag.de/de/reisefuehrer/deutschland/berlin-city/gps-daten/
# GeoJSON file will be downloaded from datahub.io/core/geo-countries
radrouten_path = geodata_path / "radrouten-berlin"
single_gpx = geodata_path / "michael-mueller-verlag-berlin.gpx"
countries = geodata_path / "countries.geojson"

# The three downloads are checked one at a time, so a missing one does not fetch the other two again
if any(radrouten_path.glob("*.gpx")) and not force:
    print(f"Keeping radrouten-berlin with {len(list(radrouten_path.glob('*.gpx')))} GPX files")
else:
    # Start from a clean directory, so a renamed file in the archive does not linger
    if radrouten_path.exists():
        shutil.rmtree(radrouten_path)
    radrouten_path.mkdir()

    print("Downloading GPX data from berlin.de")

    archive = radrouten_path / "tmp.7z"
    download(
        "https://www.berlin.de/sen/uvk/_assets/verkehr/verkehrsplanung/radverkehr/radrouten/radrouten_komplett.7z",
        archive
    )

    subprocess.run(["7za", "e", "-y", archive.name], cwd=radrouten_path, check=True, capture_output=True)

    archive.unlink()

    print(f"Created radrouten-berlin with {len(list(radrouten_path.glob('*.gpx')))} GPX files")

if single_gpx.exists() and not force:
    print(f"Keeping {single_gpx.name} with {single_gpx.stat().st_size / 1024:.0f} KB")
else:
    print("Downloading GPX data from michael-mueller-verlag.de")

    download("https://mmv.me/52630/00.gpx", single_gpx)

    print(f"Created {single_gpx.name} with {single_gpx.stat().st_size / 1024:.0f} KB")

if countries.exists() and not force:
    print(f"Keeping {countries.name} with {countries.stat().st_size / 1024 / 1024:.1f} MB")
else:
    print("Downloading GeoJSON data from datahub.io")

    download("https://datahub.io/core/geo-countries/r/0.geojson", countries)

    print(f"Created {countries.name} with {countries.stat().st_size / 1024 / 1024:.1f} MB")


# ProjectStatus
# One Excel file will be generated from sample.json
print("Setting up the Excel file for ProjectStatus")

# The demo writes ProjectStatus_Failures.xlsx into the same directory, so start from a clean one
for file in projectstatus_path.glob("*.xlsx"):
    file.unlink()

sample_data = json.loads((projectstatus_path / "sample.json").read_text(encoding="utf-8"))

# The columns are the properties of the first project, in the order they are written
columns = list(sample_data[0].keys())

workbook = Workbook()
worksheet = workbook.active
worksheet.title = "ProjectStatus"

# The note that the project managers are supposed to read
worksheet["A1"] = "Please fill out this form weekly and send it to project management office. Thanks!"
worksheet["A1"].font = Font(size=14, bold=True)

# The header row - the demo skips the first two rows and reads this one
for column, header in enumerate(columns, start=1):
    worksheet.cell(row=3, column=column, value=header).font = Font(bold=True)

# One row per project. The values go in exactly as sample.json has them, and that is the whole
# point of this scenario: "Late july 2026" is text in a column of dates and "unknown" is text in
# a column of numbers, and the database is what finally objects to them.
for index, row in enumerate(sample_data, start=4):
    for column, header in enumerate(columns, start=1):
        cell = worksheet.cell(row=index, column=column, value=row[header])

        # ProgressPercent holds "unknown" as well as numbers, and a mixed column reads better
        # lined up on one side. EPPlus lets the sibling style the whole column at once;
        # openpyxl has no column style, so the alignment is set per cell.
        if header == "ProgressPercent":
            cell.alignment = Alignment(horizontal="left")

# Make the columns wide enough to read
for column, width in {"A": 25, "B": 10, "C": 15, "D": 25, "E": 10, "F": 15, "G": 25, "H": 15}.items():
    worksheet.column_dimensions[column].width = width

file = projectstatus_path / "ProjectStatus.xlsx"
workbook.save(file)
print(f"Created {file.name} with {len(sample_data)} rows")

print("Finished")
